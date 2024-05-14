import os
import time
import random
import openpyxl

from bs4 import BeautifulSoup

from selenium import webdriver
from selenium.webdriver.common.by import By

def _driver_init():
    options = webdriver.ChromeOptions()    
    options.add_experimental_option("excludeSwitches", ["enable-logging"])
    options.add_argument("disable-gpu")
    options.add_experimental_option('useAutomationExtension', False)

    driver = webdriver.Chrome(options=options)

    driver.implicitly_wait(8)
    driver.maximize_window()
    return driver

def _custom_sleep(min=1, max=2.5):
    time.sleep(random.uniform(min, max))

def _save_excel(result, tap_name:str, filename:str):
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    for char in [":", "\\", "/", "?", "*", "[", "]"]:
        tap_name = tap_name.replace(char, "_")
    if tap_name not in wb.sheetnames:
        ws = wb.create_sheet(tap_name)
    else:
        ws = wb[tap_name]
        
    ws.append(['제목', '가격1', '가격2', '가격3', '가격4', '가격5', '링크'])
    for r in result:
        ws.append(r)

    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(filename)
    wb.close()

def temp_text(text:str):
    return text.replace('\n', '').replace(',', '').replace('원', '').replace('최저', '').replace('\xa0', '')

def main():
    products_count = 90
    urls = [
        'https://search.shopping.naver.com/best/category/click?categoryCategoryId=50000002&categoryDemo=M03&categoryRootCategoryId=50000002&period=P1D&tr=nwbhi',
        'https://search.shopping.naver.com/best/category/click?categoryCategoryId=50000003&categoryDemo=M03&categoryRootCategoryId=50000003&period=P1D&tr=nwbhi',
        'https://search.shopping.naver.com/best/category/click?categoryCategoryId=50000004&categoryDemo=M03&categoryRootCategoryId=50000004&period=P1D&tr=nwbhi',
        'https://search.shopping.naver.com/best/category/click?categoryCategoryId=50000005&categoryDemo=M03&categoryRootCategoryId=50000005&period=P1D&tr=nwbhi',
        'https://search.shopping.naver.com/best/category/click?categoryCategoryId=50000006&categoryDemo=M03&categoryRootCategoryId=50000006&period=P1D&tr=nwbhi',
        'https://search.shopping.naver.com/best/category/click?categoryCategoryId=50000007&categoryDemo=M03&categoryRootCategoryId=50000007&period=P1D&tr=nwbhi',
        'https://search.shopping.naver.com/best/category/click?categoryCategoryId=50000009&categoryDemo=M03&categoryRootCategoryId=50000009&period=P1D&tr=nwbhi',
    ]
    tap_names = [
        '화장품_미용',
        '디지털_가전',
        '가구_인테리어',
        '출산_육아',
        '식품',
        '스포츠_레저',
        '생활_건강',
    ]
    filename = f'_{time.strftime("%Y%m%d_%H%M%S")}.xlsx'
    # 결과모음 폴더 생성
    try:
        os.mkdir('결과모음')
    except FileExistsError:
        pass

    driver = _driver_init()
    for url, tap_name in zip(urls, tap_names):
        driver.get(url)
        cetegories = driver.find_elements(By.CLASS_NAME, "detailFilter_item_detail__iPrD7")
        for c in cetegories[1:]:
            rtn = []
            temp_tap_name = tap_name+ "_" + c.text
            try:
                c.click()
            except:
                driver.find_element(By.CLASS_NAME, "detailFilter_btn_next__7wfaZ").click()
                _custom_sleep(2, 3)
                c.click()
            _custom_sleep(2, 3)
            # 스크롤 다운
            for _ in range(6):
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                _custom_sleep(2, 3)

            products = driver.find_elements(By.CLASS_NAME, "imageProduct_item__KZB_F")[:products_count]
            for p in products:
                soup = BeautifulSoup(p.get_attribute("innerHTML"), 'html.parser')
                # 판매처가 없으면 패스
                if soup.find('a', class_='imageProduct_btn_store__bL4eB linkAnchor') is None:
                    continue
                p.click()
                _custom_sleep(2, 3)
                driver.switch_to.window(driver.window_handles[1])
                soup = BeautifulSoup(driver.page_source, 'html.parser')
                costs = []
                if soup.find('div', class_='top_summary_title__ViyrM') is None:
                    driver.close()
                    driver.switch_to.window(driver.window_handles[0])
                    continue
                title = soup.find('div', class_='top_summary_title__ViyrM').find('h2').text
                table = soup.find('table', class_='productByMall_list_seller__yNhgM productByMall_price_blue__wqrME')
                if table is None:
                    driver.close()
                    driver.switch_to.window(driver.window_handles[0])
                    continue
                cost_list = table.find_all('tr')
                is_official = False
                for c in cost_list[1:6]:
                    if "공식" in c.text:
                        is_official = True
                        break
                    else:
                        cost = c.find('td', class_='productByMall_price__MjaUK').text
                        costs.append(temp_text(cost))
                if is_official:
                    driver.close()
                    driver.switch_to.window(driver.window_handles[0])
                    _custom_sleep(2, 3)
                    continue
                while len(costs) < 5:
                    costs.append(None)
                rtn.append([title]+costs+[driver.current_url])
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
                _custom_sleep(2, 3)
            _save_excel(rtn, temp_tap_name, './결과모음/' + tap_name+filename)
    driver.quit()

if __name__ == "__main__":
    main()